from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


class WeChatBillParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedWeChatTransaction:
    transaction_time: str
    transaction_date: str
    transaction_type: str | None
    counterparty: str | None
    product: str | None
    direction: str | None
    amount: float
    currency: str
    payment_method: str | None
    status: str | None
    wechat_order_id: str
    merchant_order_id: str | None
    remark: str | None


_HEADER_MARKERS = ("交易时间", "记账时间")
_PERSONAL_COLUMNS = {
    "交易时间": "transaction_time",
    "交易类型": "transaction_type",
    "交易对方": "counterparty",
    "商品": "product",
    "收/支": "direction",
    "金额(元)": "amount",
    "支付方式": "payment_method",
    "当前状态": "status",
    "交易单号": "wechat_order_id",
    "商户单号": "merchant_order_id",
    "备注": "remark",
}
_FUND_COLUMNS = {
    "记账时间": "transaction_time",
    "业务名称": "transaction_type",
    "业务类型": "product",
    "收支类型": "direction",
    "收支金额(元)": "amount",
    "备注": "remark",
    "微信支付业务单号": "wechat_order_id",
}


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _clean_cell(value: Any) -> str:
    text = _cell_to_str(value)
    if text.startswith("`"):
        text = text[1:].strip()
    return text.strip()


def _parse_amount(raw: Any) -> float:
    if isinstance(raw, (int, float)):
        return abs(float(raw))
    text = _clean_cell(raw)
    text = text.replace("¥", "").replace("￥", "").replace(",", "").strip()
    if not text:
        raise WeChatBillParseError("Missing amount")
    return abs(float(text))


def _parse_datetime(raw: Any) -> tuple[str, str]:
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d %H:%M:%S"), raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat(), raw.isoformat()

    text = _clean_cell(raw)
    if not text:
        raise WeChatBillParseError("Missing transaction time")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime("%Y-%m-%d %H:%M:%S"), dt.date().isoformat()
        except ValueError:
            continue
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        date_part = text[:10]
        return text, date_part
    raise WeChatBillParseError(f"Unrecognized datetime format: {raw!r}")


def _find_header_row(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows):
        if not row:
            continue
        cleaned = [_clean_cell(cell) for cell in row]
        first = cleaned[0] if cleaned else ""
        if first in _HEADER_MARKERS:
            return idx, cleaned
    raise WeChatBillParseError(
        "Could not find WeChat bill header row (expected columns starting with 交易时间 or 记账时间)",
    )


def _column_map(headers: list[str]) -> dict[str, str]:
    if headers[0] == "交易时间":
        mapping = _PERSONAL_COLUMNS
    elif headers[0] == "记账时间":
        mapping = _FUND_COLUMNS
    else:
        raise WeChatBillParseError(f"Unsupported WeChat bill header: {headers[0]!r}")

    result: dict[str, str] = {}
    for idx, header in enumerate(headers):
        key = mapping.get(header)
        if key:
            result[key] = str(idx)
    if "transaction_time" not in result or "amount" not in result:
        raise WeChatBillParseError("WeChat bill header is missing required columns")
    return result


def _row_value(row: list[Any], col_map: dict[str, str], field: str) -> str:
    idx = col_map.get(field)
    if idx is None:
        return ""
    index = int(idx)
    if index >= len(row):
        return ""
    return _clean_cell(row[index])


def _parse_rows(rows: list[list[Any]]) -> list[ParsedWeChatTransaction]:
    header_idx, headers = _find_header_row(rows)
    col_map = _column_map(headers)

    transactions: list[ParsedWeChatTransaction] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(not _clean_cell(cell) for cell in row):
            continue
        first = _clean_cell(row[0])
        if first.startswith("-") or first.startswith("---") or "共" in first:
            break

        try:
            tx_time_raw = row[int(col_map["transaction_time"])]
            tx_time, tx_date = _parse_datetime(tx_time_raw)
            amount = _parse_amount(row[int(col_map["amount"])])
        except (WeChatBillParseError, ValueError, IndexError, KeyError):
            continue

        order_id = _row_value(row, col_map, "wechat_order_id")
        if not order_id:
            order_id = f"{tx_time}-{amount}-{len(transactions)}"

        transactions.append(
            ParsedWeChatTransaction(
                transaction_time=tx_time,
                transaction_date=tx_date,
                transaction_type=_row_value(row, col_map, "transaction_type") or None,
                counterparty=_row_value(row, col_map, "counterparty") or None,
                product=_row_value(row, col_map, "product") or None,
                direction=_row_value(row, col_map, "direction") or None,
                amount=amount,
                currency="CNY",
                payment_method=_row_value(row, col_map, "payment_method") or None,
                status=_row_value(row, col_map, "status") or None,
                wechat_order_id=order_id,
                merchant_order_id=_row_value(row, col_map, "merchant_order_id") or None,
                remark=_row_value(row, col_map, "remark") or None,
            ),
        )

    if not transactions:
        raise WeChatBillParseError("No transactions found in WeChat bill file")
    return transactions


def _rows_from_csv_text(text: str) -> list[list[Any]]:
    lines = text.splitlines()
    header_idx = None
    for idx, line in enumerate(lines):
        if any(marker in line for marker in _HEADER_MARKERS):
            header_idx = idx
            break
    if header_idx is None:
        raise WeChatBillParseError(
            "Could not find WeChat bill header row (expected columns starting with 交易时间 or 记账时间)",
        )

    rows: list[list[Any]] = []
    reader = csv.reader(io.StringIO("\n".join(lines[header_idx:])))
    for row in reader:
        rows.append(row)
    return rows


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise WeChatBillParseError("WeChat bill workbook has no active sheet")
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _is_xlsx(content: bytes, filename: str | None) -> bool:
    if filename and filename.lower().endswith(".xlsx"):
        return True
    return content[:2] == b"PK"


def parse_wechat_bill(content: bytes, *, filename: str | None = None) -> list[ParsedWeChatTransaction]:
    """Parse a WeChat Pay personal bill export (XLSX or CSV)."""
    if _is_xlsx(content, filename):
        rows = _rows_from_xlsx(content)
    else:
        text = content.decode("utf-8-sig", errors="replace").lstrip("\ufeff")
        rows = _rows_from_csv_text(text)
    return _parse_rows(rows)
